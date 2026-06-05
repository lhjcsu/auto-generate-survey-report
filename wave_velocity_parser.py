"""
波速数据解析器 — 方案二: 从实测波速数据提取

支持两种数据源:
  1. 波速数据.xlsx — 华宁全自动波速测试导出格式
     - <TC> 模板: 土层编号→土层波速映射
     - <ZK> 逐孔: 土层编号 + 土层深度
  2. 波速报告.docx — 正式波速测试报告汇总表
     - 表格含: 孔号 | 覆盖层厚度 | 计算深度d0 | νse | 场地类别

输出统一格式: [{bh_id, νse, cover_thickness, site_class, layers: [{name, vs, thickness, depth}]}]
"""

from __future__ import annotations

import os
import re
from typing import Any, Dict, List, Optional, Tuple

# ============================================================
# 数据源 1: 华宁波速数据 Excel
# ============================================================


class HuaNingWaveVelocityParser:
    """解析华宁"全自动波速测试"导出的 Excel"""

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self._tc_vs_map: Dict[str, float] = {}
        self._tc_names: Dict[str, str] = {}

    def _load_workbook(self):
        """延迟加载 openpyxl"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要 openpyxl: pip install openpyxl")
        return openpyxl.load_workbook(self.xlsx_path, data_only=True)

    def parse(self) -> List[Dict[str, Any]]:
        """解析所有钻孔波速数据

        Returns:
            [{bh_id, νse, cover_thickness, site_class, layers}]
            layers: [{layer_id, name, vs, thickness, depth}]
        """
        from wave_velocity_estimate import (
            compute_equivalent_vs,
            compute_cover_thickness,
            classify_site,
        )

        wb = self._load_workbook()
        ws = wb["全自动波速测试"]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

        # ---- Step 1: 解析模板 ----
        tc_vs_map: Dict[str, float] = {}
        tc_names: Dict[str, str] = {}
        for ri, row in enumerate(rows):
            r0 = str(row[0] or ""); r1 = str(row[1] or "") if len(row) > 1 else ""
            if r0 == "钻孔标识" and r1 == "<TC>":
                lids = rows[ri + 1]; lnames = rows[ri + 2]; vs_row = rows[ri + 4]
                for ci in range(1, 10):
                    lid = str(lids[ci] or "") if ci < len(lids) else ""
                    vs_str = str(vs_row[ci] or "") if ci < len(vs_row) else ""
                    if lid and vs_str:
                        tc_vs_map[lid] = float(vs_str.split(",")[0])
                        tc_names[lid] = str(lnames[ci] or "") if ci < len(lnames) else ""
                self._tc_vs_map = tc_vs_map
                self._tc_names = tc_names
                break

        # ---- Step 2: 解析逐孔数据 ----
        results: List[Dict[str, Any]] = []
        for ri, row in enumerate(rows):
            r0 = str(row[0] or ""); r1 = str(row[1] or "") if len(row) > 1 else ""
            if r0 == "钻孔标识" and r1 == "<ZK>" and ri + 3 < len(rows):
                info = rows[ri + 1]; lids_row = rows[ri + 2]; depths_row = rows[ri + 3]
                bh_id = str(info[3] or "").replace("ZK", "")

                layers = []; prev = 0.0
                for ci in range(1, 10):
                    lid = str(lids_row[ci] or "") if ci < len(lids_row) else ""
                    ds = str(depths_row[ci] or "") if ci < len(depths_row) else ""
                    if not lid or not ds:
                        break
                    try:
                        d = float(ds)
                    except ValueError:
                        break
                    vs = tc_vs_map.get(lid, 0)
                    layers.append({
                        "layer_id": lid,
                        "name": tc_names.get(lid, lid),
                        "vs": vs,
                        "thickness": round(d - prev, 1),
                        "depth": d,
                    })
                    prev = d

                if layers:
                    vse = compute_equivalent_vs(layers)
                    cover = compute_cover_thickness(layers)
                    site_class, _ = classify_site(vse, cover)
                    results.append({
                        "bh_id": bh_id,
                        "vse": vse,
                        "cover_thickness": cover,
                        "site_class": site_class,
                        "layers": layers,
                    })

        return results


# ============================================================
# 数据源 2: 波速报告 DOCX 汇总表
# ============================================================


class WaveVelocityReportParser:
    """解析波速测试报告 DOCX 中的汇总表"""

    def __init__(self, docx_path: str):
        self.docx_path = docx_path

    def parse_summary_table(self) -> List[Dict[str, Any]]:
        """从波速报告 DOCX 提取汇总表数据

        查找含"等效剪切波速νse"和"场地类别"的表格,
        解析 孔号 | 覆盖层厚度 | d0 | νse | 场地类别

        Returns:
            [{bh_id, vse, cover_thickness, d0, site_class}]
        """
        try:
            from docx import Document
        except ImportError:
            raise ImportError("需要 python-docx: pip install python-docx")

        doc = Document(self.docx_path)

        for tbl in doc.tables:
            # 检查表头
            row0_text = " ".join(c.text.strip() for c in tbl.rows[0].cells) if tbl.rows else ""
            if "等效剪切波速" not in row0_text or "场地类别" not in row0_text:
                continue

            # 确定列映射
            header_row = None
            for ri in range(min(3, len(tbl.rows))):
                hdr = " ".join(c.text.strip() for c in tbl.rows[ri].cells)
                if "孔号" in hdr or "序号" in hdr:
                    header_row = ri
                    break
            if header_row is None:
                continue

            results = []
            for ri in range(header_row + 1, len(tbl.rows)):
                cells = [c.text.strip() for c in tbl.rows[ri].cells]
                if not cells or all(c == "" for c in cells[1:] if len(cells) > 1):
                    continue

                # 列序: 序号 | 孔号 | 覆盖层厚度 | d0 | νse | 场地类别
                bh_id = cells[1] if len(cells) > 1 else ""
                cover_str = cells[2] if len(cells) > 2 else ""
                d0_str = cells[3] if len(cells) > 3 else ""
                vse_str = cells[4] if len(cells) > 4 else ""
                site_class = cells[5] if len(cells) > 5 else ""

                if not bh_id:
                    continue

                # 解析数值
                cover = self._parse_value(cover_str)
                d0 = self._parse_value(d0_str)
                vse = self._parse_value(vse_str)

                results.append({
                    "bh_id": bh_id.replace("ZK", ""),
                    "vse": vse,
                    "cover_thickness": cover,
                    "d0": d0,
                    "site_class": site_class.replace(" ", ""),
                })

            return results

        return []

    @staticmethod
    def _parse_value(s: str) -> Optional[float]:
        """从字符串提取数值"""
        s = s.strip().replace(" ", "")
        match = re.search(r"[\d.]+", s)
        return float(match.group()) if match else None


# ============================================================
# 统一入口
# ============================================================


def load_wave_velocity_data(
    xlsx_path: Optional[str] = None,
    docx_path: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """加载波速数据，优先报告 DOCX，回退 Excel

    Returns:
        [{bh_id, vse, cover_thickness, d0?, site_class, layers?}]
    """
    if docx_path and os.path.isfile(docx_path):
        parser = WaveVelocityReportParser(docx_path)
        results = parser.parse_summary_table()
        if results:
            return results

    if xlsx_path and os.path.isfile(xlsx_path):
        parser = HuaNingWaveVelocityParser(xlsx_path)
        return parser.parse()

    return []


# ============================================================
# 自测
# ============================================================

if __name__ == "__main__":
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

    test_xlsx = r"D:\项目\城投文化中路住宅项目\波速\波速数据.xlsx"
    test_docx = r"D:\项目\城投文化中路住宅项目\波速\波速报告（城投文化中路住宅项目）.docx"

    if os.path.isfile(test_docx):
        print("=== 波速报告 DOCX 汇总表 ===")
        rpt = WaveVelocityReportParser(test_docx)
        data = rpt.parse_summary_table()
        print(f"共 {len(data)} 个钻孔")
        for d in data[:5]:
            print(f"  ZK{d['bh_id']}: νse={d['vse']}m/s, d={d['cover_thickness']}m, {d['site_class']}类")
        if len(data) > 5:
            print(f"  ... 还有 {len(data) - 5} 个")

    if os.path.isfile(test_xlsx):
        print("\n=== 华宁波速数据 Excel ===")
        hn = HuaNingWaveVelocityParser(test_xlsx)
        data = hn.parse()
        print(f"共 {len(data)} 个钻孔")
        for d in data[:5]:
            print(f"  ZK{d['bh_id']}: νse={d['vse']}m/s, d={d['cover_thickness']}m, {d['site_class']}类")
